"""
render_sheet の run() を FakeDriveAdapter でテストする（Secrets 不要）。
CSV 取得 → テンプレ流し込み → アップロードまでのロジックを検証する。
"""
from __future__ import annotations

import io
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName

# プロジェクトルートを path に追加（render_sheet が scripts を import するため）
_repo_root = Path(__file__).resolve().parent.parent
import sys
if str(_repo_root) not in sys.path:
    sys.path.insert(0, str(_repo_root))

from scripts.gdrive.drive_client import FakeDriveAdapter
from scripts.render_sheet.render_sheet import load_config, run


def _get_xlsx_from_fake(fake: FakeDriveAdapter) -> bytes:
    """FakeDriveAdapter からアップロードされた XLSX の bytes を取得する。"""
    for meta in fake._files.values():
        name = meta.get("name", "")
        if name.endswith("Daily.xlsx"):
            content = meta["content"]
            return content if isinstance(content, bytes) else content.encode()
    raise AssertionError("XLSX が Fake に見つかりません")


def test_run_with_fake_drive_adapter() -> None:
    """FakeDriveAdapter を渡して run() が完了し、URL を返すことを検証する。"""
    config = load_config(_repo_root / "config" / "render_sheet.yaml")
    template_path_str = config.get("template_path")
    if not template_path_str:
        pytest.skip("config/render_sheet.yaml に template_path が設定されていません")
    template_path = _repo_root / template_path_str
    if not template_path.exists():
        pytest.skip(f"テンプレートがありません: {template_path_str}")

    csv_content = (
        "code,name,z_turnover_60\n"
        "7203,Toyota,0.5\n"
        "9984,SoftBank,-0.2\n"
    ).encode("utf-8-sig")
    # extract_file_id は 20〜50 文字の ID を要求する
    fake_csv_id = "fake-csv-id-12345678901234567"
    fake = FakeDriveAdapter()
    fake.put_file(fake_csv_id, csv_content, "indicators_20260222.csv")

    cfg = {
        "csv_drive_file_id": fake_csv_id,
        "output_folder_id": "fake-folder-id",
        "output_subfolder": None,
        "header_anchor_sheet_name": "indicators001",
        "template_path": template_path_str,
        "link_label_map": {},
        "sort_column": "z_turnover_60",
        "sort_ascending": False,
    }

    url = run(cfg, drive_adapter=fake)

    assert isinstance(url, str)
    assert "drive.google.com" in url or "file/d/" in url
    # アップロードが呼ばれていれば Fake にファイルが増えている（2件: CSV 登録 + XLSX アップロード）
    assert len(fake._files) >= 2


def test_sheet_protection_enabled() -> None:
    """sheet_protection: True のとき、全シートに保護がかかり、ソート・フィルタ許可、データ範囲は unlock。"""
    config = load_config(_repo_root / "config" / "render_sheet.yaml")
    template_path_str = config.get("template_path")
    if not template_path_str:
        pytest.skip("config/render_sheet.yaml に template_path が設定されていません")
    template_path = _repo_root / template_path_str
    if not template_path.exists():
        pytest.skip(f"テンプレートがありません: {template_path_str}")

    csv_content = (
        "code,name,z_turnover_60\n"
        "7203,Toyota,0.5\n"
        "9984,SoftBank,-0.2\n"
    ).encode("utf-8-sig")
    fake_csv_id = "fake-csv-id-12345678901234567"
    fake = FakeDriveAdapter()
    fake.put_file(fake_csv_id, csv_content, "indicators_20260222.csv")

    cfg = {
        "csv_drive_file_id": fake_csv_id,
        "output_folder_id": "fake-folder-id",
        "output_subfolder": None,
        "header_anchor_sheet_name": "indicators001",
        "template_path": template_path_str,
        "link_label_map": {},
        "sort_column": "z_turnover_60",
        "sort_ascending": False,
        "sheet_protection": True,
    }

    run(cfg, drive_adapter=fake)
    xlsx_bytes = _get_xlsx_from_fake(fake)
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=False)

    for ws in wb.worksheets:
        assert ws.protection.sheet is True
        assert ws.protection.sort is False
        assert ws.protection.autoFilter is False

    ws = wb["indicators001"]
    # データ範囲のセルが unlock であること（code=7203 を含むセルを探す）
    unlocked_found = False
    for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10):
        for cell in row:
            if cell.value == "7203" or cell.value == 7203:
                assert cell.protection.locked is False
                unlocked_found = True
                break
        if unlocked_found:
            break
    assert unlocked_found, "データ範囲セル（7203）が unlock であること"


def test_sheet_protection_disabled() -> None:
    """sheet_protection: False のとき、保護がかからない。"""
    config = load_config(_repo_root / "config" / "render_sheet.yaml")
    template_path_str = config.get("template_path")
    if not template_path_str:
        pytest.skip("config/render_sheet.yaml に template_path が設定されていません")
    template_path = _repo_root / template_path_str
    if not template_path.exists():
        pytest.skip(f"テンプレートがありません: {template_path_str}")

    csv_content = (
        "code,name,z_turnover_60\n"
        "7203,Toyota,0.5\n"
    ).encode("utf-8-sig")
    fake_csv_id = "fake-csv-id-12345678901234567"
    fake = FakeDriveAdapter()
    fake.put_file(fake_csv_id, csv_content, "indicators_20260222.csv")

    cfg = {
        "csv_drive_file_id": fake_csv_id,
        "output_folder_id": "fake-folder-id",
        "output_subfolder": None,
        "header_anchor_sheet_name": "indicators001",
        "template_path": template_path_str,
        "link_label_map": {},
        "sort_column": None,
        "sort_ascending": False,
        "sheet_protection": False,
    }

    run(cfg, drive_adapter=fake)
    xlsx_bytes = _get_xlsx_from_fake(fake)
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=False)

    for ws in wb.worksheets:
        assert ws.protection.sheet is False


def test_hyperlink_source_map_uses_title_column(tmp_path: Path) -> None:
    """表示列(event_news_1_title)にURL列(event_news_1_url)をリンク埋め込みできることを検証。"""
    # テンプレートをテスト用に作成
    template_path = tmp_path / "tmp_render_template.xlsx"
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "indicators001"
    ws["A1"] = "event_news_1_title"
    wb.defined_names.add(DefinedName("headerAnchor", attr_text="'indicators001'!$A$1"))
    wb.save(template_path)

    csv_content = (
        "event_news_1_title,event_news_1_url\n"
        "決算上方修正に関するお知らせ,https://example.com/news/1\n"
        "材料不明・需給起因疑い,\n"
    ).encode("utf-8-sig")
    fake_csv_id = "fake-csv-id-12345678901234567"
    fake = FakeDriveAdapter()
    fake.put_file(fake_csv_id, csv_content, "indicators_event_enriched_20260307.csv")

    cfg = {
        "csv_drive_file_id": fake_csv_id,
        "output_folder_id": "fake-folder-id",
        "output_subfolder": None,
        "header_anchor_sheet_name": "indicators001",
        "template_path": str(template_path),
        "link_label_map": {},
        "hyperlink_source_map": {"event_news_1_title": "event_news_1_url"},
        "sort_column": None,
        "sort_ascending": False,
        "sheet_protection": False,
    }

    run(cfg, drive_adapter=fake)
    xlsx_bytes = _get_xlsx_from_fake(fake)
    out_wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=False)
    out_ws = out_wb["indicators001"]

    assert out_ws["A2"].value == "決算上方修正に関するお知らせ"
    assert out_ws["A2"].hyperlink is not None
    assert out_ws["A2"].hyperlink.target == "https://example.com/news/1"

    assert out_ws["A3"].value == "材料不明・需給起因疑い"
    assert out_ws["A3"].hyperlink is None

