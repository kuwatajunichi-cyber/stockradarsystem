"""Create indicators_template_v1.4.xlsx from v1.3."""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

_REPO = Path(__file__).resolve().parents[2]


def main() -> None:
    src = _REPO / "config" / "templates" / "indicators_template_v1.3.xlsx"
    dst = _REPO / "config" / "templates" / "indicators_template_v1.4.xlsx"
    if not src.is_file():
        print(f"missing source: {src}", file=sys.stderr)
        sys.exit(1)
    shutil.copy2(src, dst)

    wb = load_workbook(dst, data_only=False)
    dn = wb.defined_names.get("headerAnchor")
    if not dn:
        print("missing headerAnchor", file=sys.stderr)
        sys.exit(1)

    row1 = "\u89b3\u6e2c\u30c7\u30fc\u30bf\uff08\u8abf\u67fb\u30d7\u30ed\u30f3\u30d7\u30c8\u7528\uff09"
    row2 = "\u62d9\u7c8b\u30d6\u30ed\u30c3\u30af"

    for title, _coord in dn.destinations:
        ws = wb[str(title)]
        new_col = 19
        ref_col = 18
        ws.cell(row=1, column=new_col).value = row1
        ws.cell(row=2, column=new_col).value = row2
        ws.cell(row=3, column=new_col).value = "research_prompt_block"

        if ws.auto_filter and ws.auto_filter.ref:
            ws.auto_filter.ref = f"A3:{get_column_letter(new_col)}3000"

        dim = ws.column_dimensions.get(get_column_letter(ref_col))
        if dim and dim.width is not None:
            ws.column_dimensions[get_column_letter(new_col)].width = dim.width
        else:
            ws.column_dimensions[get_column_letter(new_col)].width = 36.0

    wb.save(dst)
    print(f"wrote {dst}")


if __name__ == "__main__":
    main()
