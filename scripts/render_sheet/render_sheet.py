"""
Google Drive 上の CSV を読み込み、ローカル XLSX テンプレートに流し込んで日次レポートを生成する。

- 入力: Drive 上の CSV（file ID または共有リンクで指定）
- テンプレ: ローカル XLSX（openpyxl で処理）
- 出力: XLSX を Drive にアップロード
- 認証: OAuth（GDRIVE_OAUTH_*、Drive API のみ・Sheets API 不要）
"""
from __future__ import annotations

import argparse
import io
import logging
import re
import sys
from pathlib import Path

import pandas as pd

# openpyxl はパス変更前にインポート（sys.path の影響を受けないように）
from openpyxl import load_workbook
from openpyxl.styles import Font, Protection
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from openpyxl.worksheet.protection import SheetProtection

# プロジェクトルートを PYTHONPATH に追加
_script_dir = Path(__file__).resolve().parent
_repo_root = _script_dir.parent.parent
if str(_repo_root) not in sys.path:
    sys.path.insert(0, str(_repo_root))

from scripts.gdrive.drive_client import (
    DriveAdapter,
    GoogleDriveAdapter,
    build_service,
    get_credentials,
    get_folder_id_paid,
    get_file_metadata,
    get_or_create_folder,
    upload_file,
)

logger = logging.getLogger(__name__)

# XLSX MIME type
MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# --- Drive リンク → fileId 抽出 ---
def extract_file_id(value: str) -> str:
    """Drive の file ID を抽出する。"""
    s = (value or "").strip()
    if not s:
        raise ValueError("ファイルIDまたは共有リンクが空です")
    m = re.search(r"/file/d/([a-zA-Z0-9_-]+)", s)
    if m:
        return m.group(1)
    m = re.search(r"[?&]id=([a-zA-Z0-9_-]+)", s)
    if m:
        return m.group(1)
    if re.match(r"^[a-zA-Z0-9_-]{20,50}$", s):
        return s
    raise ValueError(f"ファイルIDを抽出できませんでした: {value[:80]}...")


def extract_date_from_filename(name: str) -> str | None:
    """ファイル名から YYYY-MM-DD を抽出。"""
    m = re.search(r"(\d{4})(\d{2})(\d{2})", name)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", name)
    if m:
        return m.group(0)
    return None


# --- 設定読み込み ---
def load_config(config_path: Path) -> dict:
    """config YAML を読み込む"""
    try:
        import yaml
    except ImportError:
        raise SystemExit("PyYAML がインストールされていません。pip install pyyaml を実行してください。")

    if not config_path.exists():
        raise SystemExit(f"設定ファイルが見つかりません: {config_path}")

    with config_path.open(encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def resolve_config(
    config: dict,
    csv_drive_file_id: str | None,
    output_folder_id: str | None,
    output_subfolder: str | None,
    header_anchor_sheet_name: str | None,
    template_path: str | Path | None,
) -> dict:
    """コマンドライン・workflow inputs で上書きした最終設定を返す"""
    out = {
        "csv_drive_file_id": csv_drive_file_id or config.get("csv_drive_file_id"),
        "output_folder_id": output_folder_id or config.get("output_folder_id") or get_folder_id_paid(),
        "output_subfolder": output_subfolder or config.get("output_subfolder"),
        # デフォルトは headerAnchor が定義されている全シートを対象にする
        # （特定シートに限定したい場合だけ CLI で指定する）
        "header_anchor_sheet_name": (header_anchor_sheet_name or "").strip() or None,
        "template_path": template_path or config.get("template_path"),
        "link_label_map": config.get("link_label_map") or {},
        "hyperlink_source_map": config.get("hyperlink_source_map") or {},
        "sort_column": config.get("sort_column"),
        "sort_ascending": config.get("sort_ascending", False),
        "sheet_protection": config.get("sheet_protection", False),
        "sheet_protection_password": config.get("sheet_protection_password") or "",
    }
    if not out["csv_drive_file_id"]:
        raise SystemExit("csv_drive_file_id が指定されていません。")
    if not out["output_folder_id"]:
        raise SystemExit("output_folder_id が指定されていません。")
    if not out["template_path"]:
        raise SystemExit("template_path が指定されていません。config/render_sheet.yaml で設定してください。")
    return out


def _parse_header_anchors(wb) -> dict[str, tuple[int, int]]:
    """
    headerAnchor の Named Range から、
    {シート名 -> (1-based row, 1-based col)} を返す。

    README 等、headerAnchor が定義されていないシートは含まれない。
    """
    dn = wb.defined_names.get("headerAnchor")
    if not dn:
        raise SystemExit("Named Range 'headerAnchor' がテンプレートに見つかりません。")

    out: dict[str, tuple[int, int]] = {}
    for title, coord in dn.destinations:
        sheet_name = str(title)
        # coord は "A5" や "A5:E5" の形式。先頭セルを取得
        addr = coord.replace("$", "").split(":")[0]
        col_letter, row = coordinate_from_string(addr)
        col_idx = column_index_from_string(col_letter)
        out[sheet_name] = (row, col_idx)

    if not out:
        raise SystemExit("Named Range 'headerAnchor' は見つかったが、destinations が空です。")
    return out


def _read_template_headers(ws, header_row: int, header_col: int, max_cols: int = 100) -> list[str]:
    """ヘッダー行を左から読み、空セルで終了するまで取得。"""
    headers = []
    for c in range(header_col, header_col + max_cols):
        cell = ws.cell(row=header_row, column=c)
        val = cell.value
        if val is None or (isinstance(val, str) and not val.strip()):
            break
        headers.append(str(val).strip())
    return headers


# --- テンプレート流し込み（Drive 非依存）---
def _build_xlsx_from_df(cfg: dict, df: pd.DataFrame, run_date: str) -> Path:
    """
    DataFrame をテンプレートに流し込み、XLSX を保存する。
    戻り値: 保存したファイルの Path。Drive には触れない。
    """
    output_name = f"{run_date}_Daily.xlsx"
    output_dir = _repo_root / "data" / "indicators" / "daily"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / output_name

    template_path = Path(cfg["template_path"])
    if not template_path.is_absolute():
        template_path = _repo_root / template_path
    if not template_path.exists():
        raise SystemExit(f"テンプレートが見つかりません: {template_path}")

    logger.info("テンプレート: %s", template_path)
    wb = load_workbook(template_path, data_only=False)
    header_anchors = _parse_header_anchors(wb)

    restrict_sheet_name = cfg.get("header_anchor_sheet_name")
    if restrict_sheet_name:
        if restrict_sheet_name not in header_anchors:
            raise SystemExit(
                f"指定されたシート '{restrict_sheet_name}' に headerAnchor がありません。"
                f"利用可能: {sorted(header_anchors.keys())}"
            )
        target_sheet_names = [restrict_sheet_name]
    else:
        # headerAnchor が存在するシートだけを対象にする（README 等は無視）
        target_sheet_names = sorted(header_anchors.keys())

    logger.info("headerAnchor 対象シート: %s", target_sheet_names)

    def _to_value(val):
        if pd.isna(val):
            return ""
        if isinstance(val, (int, float)):
            return val
        return str(val)

    link_label_map = cfg.get("link_label_map") or {}
    hyperlink_source_map = cfg.get("hyperlink_source_map") or {}
    hyperlink_font = Font(u="single", color="0563C1")

    # sheet_protection は「全シートに適用」する一方で、unlock 範囲は
    # headerAnchor 対象シートだけに限定する。
    sp: SheetProtection | None = None
    if cfg.get("sheet_protection"):
        sp = SheetProtection(
            sheet=True,
            sort=False,
            autoFilter=False,
            formatCells=True,
            insertRows=True,
            insertColumns=True,
            deleteRows=True,
            deleteColumns=True,
            insertHyperlinks=True,
            pivotTables=True,
            selectLockedCells=False,
            selectUnlockedCells=False,
        )
        if cfg.get("sheet_protection_password"):
            sp.set_password(cfg["sheet_protection_password"])

    for target_sheet_name in target_sheet_names:
        ws = wb[target_sheet_name]
        header_row, header_col = header_anchors[target_sheet_name]
        template_headers = _read_template_headers(ws, header_row, header_col)
        logger.info(
            "[%s] テンプレ列数: %d, ヘッダー: %s",
            target_sheet_name,
            len(template_headers),
            template_headers[:5],
        )

        csv_columns = set(df.columns)
        template_set = set(template_headers)
        missing_in_csv = template_set - csv_columns
        extra_in_csv = csv_columns - template_set
        if missing_in_csv:
            logger.info("[%s] テンプレに存在するがCSVに無い列（空欄）: %s", target_sheet_name, sorted(missing_in_csv))
        if extra_in_csv:
            logger.info("[%s] CSVに存在するがテンプレに無い列（無視）: %s", target_sheet_name, sorted(extra_in_csv))

        data_start_row = header_row + 1

        for row_idx, (_, r) in enumerate(df.iterrows()):
            excel_row = data_start_row + row_idx
            for col_idx, h in enumerate(template_headers):
                cell = ws.cell(row=excel_row, column=header_col + col_idx)
                if h in df.columns:
                    val = r[h]
                    if h in hyperlink_source_map:
                        display_text = _to_value(val)
                        url_col = str(hyperlink_source_map[h]).strip()
                        url_val = r[url_col] if url_col and url_col in df.columns else ""
                        if (
                            pd.notna(url_val)
                            and str(url_val).strip().startswith("http")
                            and str(display_text).strip()
                        ):
                            cell.value = display_text
                            cell.hyperlink = str(url_val).strip()
                            cell.font = hyperlink_font
                        else:
                            cell.value = display_text
                    elif h in link_label_map:
                        label = link_label_map[h]
                        if pd.notna(val) and str(val).strip().startswith("http"):
                            url_str = str(val).strip()
                            cell.value = label
                            cell.hyperlink = url_str
                            cell.font = hyperlink_font
                        else:
                            cell.value = _to_value(val)
                    else:
                        cell.value = _to_value(val)
                else:
                    cell.value = ""

        last_row = data_start_row + len(df) - 1
        logger.info("[%s] 書き込み最終行: %d (1-based)", target_sheet_name, last_row)

        format_threshold = 3000
        if last_row > format_threshold:
            logger.warning("[%s] 書式設定済み範囲（約%d行）を超えて書き込みました。", target_sheet_name, format_threshold)
        else:
            logger.info("[%s] 書式設定済み範囲内で書き込みました。", target_sheet_name)

        label_url_cols = [c for c in template_headers if c in link_label_map]
        dynamic_cols = [c for c in template_headers if c in hyperlink_source_map]
        url_count_label = sum(
            1
            for _, r in df.iterrows()
            for c in label_url_cols
            if c in r and pd.notna(r.get(c)) and str(r.get(c)).startswith("http")
        )
        url_count_dynamic = sum(
            1
            for _, r in df.iterrows()
            for c in dynamic_cols
            if c in r
            and str(r.get(c)).strip()
            and str(hyperlink_source_map.get(c, "")).strip() in r.index
            and pd.notna(r.get(str(hyperlink_source_map.get(c, "")).strip()))
            and str(r.get(str(hyperlink_source_map.get(c, "")).strip())).startswith("http")
        )
        logger.info(
            "[%s] URL置換件数: fixed=%d dynamic=%d (fixed列: %s, dynamic列: %s)",
            target_sheet_name,
            url_count_label,
            url_count_dynamic,
            label_url_cols,
            dynamic_cols,
        )

        if sp:
            end_col = header_col + len(template_headers) - 1
            for row in range(header_row, last_row + 1):
                for col in range(header_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.protection = Protection(locked=False)

    if sp:
        for sheet in wb.worksheets:
            sheet.protection = sp
        logger.info("シート保護を適用しました（ソート・フィルタ許可）")

    wb.save(output_path)
    logger.info("ローカル保存: %s", output_path)
    return output_path


def run_local(input_csv: Path, cfg: dict) -> Path:
    """
    ローカル CSV を読み、XLSX テンプレに流し込んで日次レポートを生成する。
    Drive には一切アクセスしない。戻り値は生成した XLSX の Path。
    """
    df = pd.read_csv(input_csv)
    logger.info("入力CSV: %s, 行数: %d", input_csv.name, len(df))

    sort_column = cfg.get("sort_column")
    if sort_column and sort_column in df.columns:
        sort_asc = cfg.get("sort_ascending", False)
        df = df.sort_values(by=sort_column, ascending=sort_asc, na_position="last")
        logger.info("ソート適用: %s %s", sort_column, "昇順" if sort_asc else "降順")
    elif sort_column:
        logger.warning("ソートキー列 '%s' がCSVに存在しないため、ソートをスキップしました。", sort_column)

    run_date = extract_date_from_filename(input_csv.name)
    if not run_date:
        run_date = pd.Timestamp.now(tz="Asia/Tokyo").strftime("%Y-%m-%d")
        logger.warning("CSVファイル名から日付を抽出できず、本日を使用: %s", run_date)

    return _build_xlsx_from_df(cfg, df, run_date)


# --- Drive モード（将来削除予定のため --csv-drive-file-id は暫定維持）---
def run(cfg: dict, drive_adapter: DriveAdapter | None = None) -> str:
    """
    メイン処理（Drive モード）。戻り値は生成したファイルの Drive URL。
    drive_adapter 未指定時は get_credentials + build_service で本番接続する。
    テスト時は FakeDriveAdapter を渡して Secrets 不要で実行可能。
    """
    if drive_adapter is None:
        creds = get_credentials()
        service = build_service(creds)
        drive: DriveAdapter = GoogleDriveAdapter(service)
    else:
        drive = drive_adapter

    csv_file_id = extract_file_id(cfg["csv_drive_file_id"])
    try:
        resp = drive.get_file_content(csv_file_id)
    except Exception as e:
        raise SystemExit(f"CSV のダウンロードに失敗しました。file_id={csv_file_id}: {e}") from e

    meta = drive.get_file_metadata(csv_file_id)
    csv_filename = meta.get("name", "")
    logger.info("入力CSV: %s (file_id=%s)", csv_filename, csv_file_id)

    df = pd.read_csv(io.BytesIO(resp))
    logger.info("入力CSV行数: %d", len(df))

    sort_column = cfg.get("sort_column")
    if sort_column and sort_column in df.columns:
        sort_asc = cfg.get("sort_ascending", False)
        df = df.sort_values(by=sort_column, ascending=sort_asc, na_position="last")
        logger.info("ソート適用: %s %s", sort_column, "昇順" if sort_asc else "降順")
    elif sort_column:
        logger.warning("ソートキー列 '%s' がCSVに存在しないため、ソートをスキップしました。", sort_column)

    run_date = extract_date_from_filename(csv_filename)
    if not run_date:
        run_date = pd.Timestamp.now(tz="Asia/Tokyo").strftime("%Y-%m-%d")
        logger.warning("CSVファイル名から日付を抽出できず、本日を使用: %s", run_date)

    output_name = f"{run_date}_Daily.xlsx"
    output_path = _build_xlsx_from_df(cfg, df, run_date)

    content = output_path.read_bytes()
    parent_id = cfg["output_folder_id"]
    if cfg.get("output_subfolder"):
        parent_id = drive.get_or_create_folder(parent_id, cfg["output_subfolder"])
        logger.info("出力先サブフォルダ: %s", cfg["output_subfolder"])
    file_id, web_link = drive.upload_file(
        parent_id, output_name, content, mime_type=MIME_XLSX
    )
    result_url = web_link or f"https://drive.google.com/file/d/{file_id}/view"
    logger.info("出力URL: %s", result_url)
    return result_url


def main() -> None:
    parser = argparse.ArgumentParser(
        description="CSV を読み込み、XLSX テンプレに流し込んで日次レポートを生成（ローカル or Drive）"
    )
    parser.add_argument(
        "--input-csv",
        type=Path,
        default=None,
        help="ローカル CSV パス。指定時は Drive を使わずローカルモード。--csv-drive-file-id より優先。",
    )
    parser.add_argument(
        "--csv-drive-file-id",
        default=None,
        help="CSV の Drive ファイル ID または共有リンク（ローカルモード時は不要。将来削除予定）",
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=_repo_root / "config" / "render_sheet.yaml",
        help="設定 YAML パス",
    )
    parser.add_argument(
        "--output-folder-id",
        default=None,
        help="出力先フォルダ ID（Drive モード時のみ。省略時は config を使用）",
    )
    parser.add_argument(
        "--output-subfolder",
        default=None,
        help="出力先サブフォルダ名（Drive モード時のみ。例: YYYY-MM）",
    )
    parser.add_argument(
        "--header-anchor-sheet-name",
        default=None,
        help="特定の headerAnchor シート名に限定（省略時は headerAnchor が存在する全シートを対象）",
    )
    parser.add_argument(
        "--template-path",
        default=None,
        help="テンプレート XLSX のパス（省略時は config を使用）",
    )
    parser.add_argument(
        "-v",
        "--verbose",
        action="store_true",
        help="詳細ログ",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(message)s",
        stream=sys.stderr,
    )

    config = load_config(args.config)

    if args.input_csv is not None:
        if not args.input_csv.is_file():
            raise SystemExit(f"入力CSVが見つかりません: {args.input_csv}")
        template_path = args.template_path or config.get("template_path")
        if not template_path:
            raise SystemExit("template_path が指定されていません。config/render_sheet.yaml で設定してください。")
        cfg = {
            "template_path": str(template_path),
            "header_anchor_sheet_name": args.header_anchor_sheet_name or None,
            "link_label_map": config.get("link_label_map") or {},
            "hyperlink_source_map": config.get("hyperlink_source_map") or {},
            "sort_column": config.get("sort_column"),
            "sort_ascending": config.get("sort_ascending", False),
            "sheet_protection": config.get("sheet_protection", False),
            "sheet_protection_password": config.get("sheet_protection_password") or "",
        }
        out_path = run_local(args.input_csv, cfg)
        print(f"output_xlsx={out_path}")
        return

    if not args.csv_drive_file_id:
        raise SystemExit("--csv-drive-file-id または --input-csv のいずれかを指定してください。")
    cfg = resolve_config(
        config,
        csv_drive_file_id=args.csv_drive_file_id,
        output_folder_id=args.output_folder_id,
        output_subfolder=args.output_subfolder,
        header_anchor_sheet_name=args.header_anchor_sheet_name,
        template_path=args.template_path,
    )
    url = run(cfg)
    print(f"spreadsheet_url={url}")


if __name__ == "__main__":
    try:
        main()
    except SystemExit:
        raise
    except Exception as e:
        logger.exception("エラー: %s", e)
        sys.exit(1)
